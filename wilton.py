import streamlit as st
import pandas as pd

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from io import BytesIO
from datetime import date, timedelta
from copy import copy


# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="Loom Production Planner",
    page_icon="🏭",
    layout="wide"
)


# ============================================================
# DEFAULT LOOM SETTINGS
# ============================================================

DEFAULT_SETTINGS = pd.DataFrame([
    {"Use": True, "Loom": "ALT1", "Weekly Capacity": 1500, "Starting Week": 33},
    {"Use": True, "Loom": "ALT2", "Weekly Capacity": 1200, "Starting Week": 33},
    {"Use": True, "Loom": "ALT4", "Weekly Capacity": 800, "Starting Week": 34},
    {"Use": True, "Loom": "ALT5", "Weekly Capacity": 1300, "Starting Week": 33},
    {"Use": True, "Loom": "ALT6", "Weekly Capacity": 800, "Starting Week": 33},

    {"Use": True, "Loom": "RP1", "Weekly Capacity": 500, "Starting Week": 33},
    {"Use": True, "Loom": "RP2", "Weekly Capacity": 700, "Starting Week": 33},
    {"Use": True, "Loom": "RP3", "Weekly Capacity": 900, "Starting Week": 33},
    {"Use": True, "Loom": "RP4", "Weekly Capacity": 900, "Starting Week": 33},
    {"Use": True, "Loom": "RP5", "Weekly Capacity": 1100, "Starting Week": 33},
    {"Use": True, "Loom": "RP6", "Weekly Capacity": 1200, "Starting Week": 33},
    {"Use": True, "Loom": "RP7", "Weekly Capacity": 1000, "Starting Week": 33},

    {"Use": True, "Loom": "DB4M1", "Weekly Capacity": 300, "Starting Week": 33},
])


if "loom_settings" not in st.session_state:
    st.session_state.loom_settings = DEFAULT_SETTINGS.copy()


# ============================================================
# BASIC HELPERS
# ============================================================

def clean_loom(value):

    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .upper()
        .replace(" ", "")
        .replace("-", "")
    )


def clean_document(value):

    if value is None:
        return ""

    return str(value).strip()


def parse_quantity(value):

    if value is None:
        return None

    try:
        quantity = float(value)
    except (ValueError, TypeError):
        return None

    if quantity <= 0:
        return None

    return quantity


def next_week(week):

    if week >= 52:
        return 1

    return week + 1


# ============================================================
# COMPANY WEEK CALENDAR
# ============================================================
#
# COMPANY RULE:
#
# Week 1  = Jan 1  - Jan 7
# Week 2  = Jan 8  - Jan 14
# ...
# Week 52 = Dec 24 - Dec 30
#
# Dec 31 is outside the 52-week calendar.
#
# NO specific date such as 10/08/2026 is hard-coded.
# ============================================================


def get_week_start_date(
    production_week,
    year
):
    """
    Convert production week to its starting date.

    Week 1 starts on January 1.
    Every following week starts seven days later.
    """

    production_week = int(production_week)

    if production_week < 1 or production_week > 52:
        raise ValueError(
            "Production week must be between 1 and 52."
        )

    return (
        date(year, 1, 1)
        + timedelta(
            days=(production_week - 1) * 7
        )
    )


def get_week_end_date(
    production_week,
    year
):

    start_date = get_week_start_date(
        production_week,
        year
    )

    return (
        start_date
        + timedelta(days=6)
    )


def get_company_week_from_date(
    production_date
):
    """
    Convert date to company production week.

    Week 1 = Jan 1-7
    ...
    Week 52 = Dec 24-30

    Dec 31 is invalid/outside the 52-week calendar.
    """

    year = production_date.year

    first_day = date(
        year,
        1,
        1
    )

    last_valid_day = (
        first_day
        + timedelta(days=(52 * 7) - 1)
    )

    if production_date < first_day:
        return None

    if production_date > last_valid_day:
        return None

    difference = (
        production_date - first_day
    ).days

    return (
        difference // 7
    ) + 1


# ============================================================
# PARSE PRODUCTION WEEK
# ============================================================

def parse_week(value):

    if value is None:
        return None

    text = (
        str(value)
        .strip()
        .upper()
        .replace("WK-", "")
        .replace("WEEK", "")
        .strip()
    )

    try:

        week = int(
            float(text)
        )

    except (ValueError, TypeError):

        return None

    if week < 1 or week > 52:
        return None

    return week


# ============================================================
# PARSE EXCEL DATE
# ============================================================

def parse_excel_date(value):

    if value is None:
        return None

    try:

        parsed = pd.to_datetime(
            value,
            errors="coerce"
        )

        if pd.isna(parsed):
            return None

        return parsed.date()

    except Exception:

        return None


# ============================================================
# VALIDATE EXISTING WEEK / DATE ALLOCATION
# ============================================================

def validate_input_allocations(
    worksheet,
    loom_settings,
    document_col,
    quantity_col,
    loom_col,
    week_col,
    date_col
):

    errors = []
    warnings = []

    known_looms = {
        clean_loom(loom)
        for loom in loom_settings.keys()
    }

    enabled_looms = {
        clean_loom(loom)
        for loom, settings in loom_settings.items()
        if settings["use"]
    }

    # --------------------------------------------------------
    # Track document sequence.
    # --------------------------------------------------------

    document_occurrences = {}

    # --------------------------------------------------------
    # Track weekly quantities.
    # --------------------------------------------------------

    weekly_quantities = {}

    for row in range(
        2,
        worksheet.max_row + 1
    ):

        loom = clean_loom(
            worksheet.cell(
                row=row,
                column=loom_col
            ).value
        )

        document = clean_document(
            worksheet.cell(
                row=row,
                column=document_col
            ).value
        )

        quantity = parse_quantity(
            worksheet.cell(
                row=row,
                column=quantity_col
            ).value
        )

        week_value = worksheet.cell(
            row=row,
            column=week_col
        ).value

        date_value = worksheet.cell(
            row=row,
            column=date_col
        ).value

        # ----------------------------------------------------
        # Completely empty row.
        # ----------------------------------------------------

        if (
            loom == ""
            and document == ""
            and quantity is None
            and week_value is None
            and date_value is None
        ):
            continue

        # ----------------------------------------------------
        # Loom.
        # ----------------------------------------------------

        if loom == "":

            errors.append(
                f"Row {row}: Loom is blank."
            )

        elif loom not in known_looms:

            errors.append(
                f"Row {row}: Loom '{loom}' "
                "is not configured."
            )

        elif loom not in enabled_looms:

            warnings.append(
                f"Row {row}: Loom '{loom}' is disabled."
            )

        # ----------------------------------------------------
        # Quantity.
        # ----------------------------------------------------

        raw_quantity = worksheet.cell(
            row=row,
            column=quantity_col
        ).value

        if (
            raw_quantity is not None
            and
            quantity is None
        ):

            errors.append(
                f"Row {row}: Invalid quantity "
                f"'{raw_quantity}'."
            )

        # ----------------------------------------------------
        # Production Week.
        # ----------------------------------------------------

        week = parse_week(
            week_value
        )

        if week_value is not None:

            if week is None:

                errors.append(
                    f"Row {row}: Invalid Production Week "
                    f"'{week_value}'. "
                    "Valid values are WK-1 to WK-52."
                )

        # ----------------------------------------------------
        # Production Date.
        # ----------------------------------------------------

        parsed_date = None

        if date_value is not None:

            parsed_date = parse_excel_date(
                date_value
            )

            if parsed_date is None:

                errors.append(
                    f"Row {row}: Invalid Production Date "
                    f"'{date_value}'."
                )

        # ----------------------------------------------------
        # Check date is inside the 52-week calendar.
        # ----------------------------------------------------

        if parsed_date is not None:

            calculated_week = (
                get_company_week_from_date(
                    parsed_date
                )
            )

            if calculated_week is None:

                errors.append(
                    f"Row {row}: Date "
                    f"{parsed_date.strftime('%d-%m-%Y')} "
                    "is outside the 52-week production calendar."
                )

        # ----------------------------------------------------
        # Check week/date consistency.
        # ----------------------------------------------------

        if (
            week is not None
            and
            parsed_date is not None
        ):

            calculated_week = (
                get_company_week_from_date(
                    parsed_date
                )
            )

            if calculated_week != week:

                errors.append(
                    f"Row {row}: Production Week/date mismatch. "
                    f"WK-{week} does not match "
                    f"{parsed_date.strftime('%d-%m-%Y')}."
                )

        # ----------------------------------------------------
        # Store document occurrences.
        # ----------------------------------------------------

        if document != "":

            key = (
                loom,
                document
            )

            if key not in document_occurrences:

                document_occurrences[
                    key
                ] = []

            document_occurrences[
                key
            ].append(
                (
                    row,
                    parsed_date,
                    week
                )
            )

        # ----------------------------------------------------
        # Weekly quantity.
        # ----------------------------------------------------

        if (
            loom != ""
            and
            week is not None
            and
            quantity is not None
        ):

            key = (
                loom,
                week
            )

            weekly_quantities[
                key
            ] = (
                weekly_quantities.get(
                    key,
                    0
                )
                + quantity
            )

    # ========================================================
    # DOCUMENT SEQUENCE VALIDATION
    # ========================================================

    for (
        loom,
        document
    ), occurrences in document_occurrences.items():

        valid_occurrences = [
            x
            for x in occurrences
            if x[1] is not None
        ]

        if len(valid_occurrences) <= 1:
            continue

        dates = [
            x[1]
            for x in valid_occurrences
        ]

        unique_dates = sorted(
            set(dates)
        )

        if len(unique_dates) <= 1:
            continue

        first_row = min(
            x[0]
            for x in valid_occurrences
        )

        last_row = max(
            x[0]
            for x in valid_occurrences
        )

        documents_between = []

        for check_row in range(
            first_row,
            last_row + 1
        ):

            check_loom = clean_loom(
                worksheet.cell(
                    row=check_row,
                    column=loom_col
                ).value
            )

            check_doc = clean_document(
                worksheet.cell(
                    row=check_row,
                    column=document_col
                ).value
            )

            if (
                check_loom == loom
                and
                check_doc != ""
                and
                check_doc != document
            ):

                documents_between.append(
                    check_doc
                )

        if documents_between:

            warnings.append(
                f"Loom {loom}, External Document "
                f"{document}: the document appears again "
                "after another document. Check the sequence."
            )

    # ========================================================
    # CAPACITY VALIDATION
    # ========================================================

    for (
        loom,
        week
    ), total_quantity in weekly_quantities.items():

        settings = loom_settings.get(
            loom
        )

        if settings is None:
            continue

        capacity = settings[
            "capacity"
        ]

        if total_quantity > capacity:

            warnings.append(
                f"Loom {loom}, WK-{week}: "
                f"allocated quantity {total_quantity:g} "
                f"exceeds weekly capacity "
                f"{capacity:g}."
            )

    return errors, warnings


# ============================================================
# BALANCED DATE ASSIGNMENT
# ============================================================

def assign_balanced_dates(
    rows,
    week_start_date
):

    """
    Balance complete product rows over seven days.

    IMPORTANT:

    - No daily capacity.
    - No product row is split.
    - Original row/document sequence remains unchanged.
    - A document can continue to the next day.
    - Next document can begin on the same day.
    """

    if not rows:
        return {}

    number_of_rows = len(rows)

    quantities = [
        float(item[1])
        for item in rows
    ]

    total_quantity = sum(
        quantities
    )

    # --------------------------------------------------------
    # Fewer than 7 rows.
    # --------------------------------------------------------

    if number_of_rows <= 7:

        result = {}

        for index, item in enumerate(rows):

            excel_row = item[0]

            result[
                excel_row
            ] = (
                week_start_date
                + timedelta(days=index)
            )

        return result

    # --------------------------------------------------------
    # Ideal quantity.
    #
    # This is NOT a daily capacity.
    # --------------------------------------------------------

    ideal = (
        total_quantity / 7
    )

    # --------------------------------------------------------
    # Prefix sums.
    # --------------------------------------------------------

    prefix = [0.0]

    for quantity in quantities:

        prefix.append(
            prefix[-1] + quantity
        )

    INF = float("inf")

    dp = [
        [INF] * (number_of_rows + 1)
        for _ in range(8)
    ]

    parent = [
        [None] * (number_of_rows + 1)
        for _ in range(8)
    ]

    dp[0][0] = 0.0

    # --------------------------------------------------------
    # Dynamic programming.
    # --------------------------------------------------------

    for day_count in range(
        1,
        8
    ):

        for end in range(
            day_count,
            number_of_rows + 1
        ):

            for start in range(
                day_count - 1,
                end
            ):

                previous = dp[
                    day_count - 1
                ][start]

                if previous == INF:
                    continue

                day_quantity = (
                    prefix[end]
                    - prefix[start]
                )

                difference = (
                    day_quantity
                    - ideal
                )

                cost = (
                    difference
                    * difference
                )

                total_cost = (
                    previous
                    + cost
                )

                if total_cost < dp[
                    day_count
                ][end]:

                    dp[
                        day_count
                    ][end] = total_cost

                    parent[
                        day_count
                    ][end] = start

    # --------------------------------------------------------
    # Recover groups.
    # --------------------------------------------------------

    days_to_use = min(
        7,
        number_of_rows
    )

    boundaries = []

    end = number_of_rows

    for day_count in range(
        days_to_use,
        0,
        -1
    ):

        start = parent[
            day_count
        ][end]

        if start is None:
            break

        boundaries.append(
            (
                start,
                end
            )
        )

        end = start

    boundaries.reverse()

    # --------------------------------------------------------
    # Assign dates.
    # --------------------------------------------------------

    result = {}

    for day_index, (
        start,
        end
    ) in enumerate(boundaries):

        assigned_date = (
            week_start_date
            + timedelta(
                days=day_index
            )
        )

        for index in range(
            start,
            end
        ):

            excel_row = rows[
                index
            ][0]

            result[
                excel_row
            ] = assigned_date

    return result


# ============================================================
# SORT FINAL EXCEL
# ============================================================

def sort_planned_rows(
    worksheet,
    loom_settings,
    loom_col,
    week_col,
    date_col,
    document_col,
    row_sequence
):

    merged_ranges = [
        str(rng)
        for rng in worksheet.merged_cells.ranges
    ]

    # --------------------------------------------------------
    # Temporarily unmerge.
    # --------------------------------------------------------

    for rng in list(
        worksheet.merged_cells.ranges
    ):

        try:

            worksheet.unmerge_cells(
                str(rng)
            )

        except Exception:

            pass

    # --------------------------------------------------------
    # Loom order.
    # --------------------------------------------------------

    loom_order = {}

    for index, loom in enumerate(
        loom_settings.keys()
    ):

        loom_order[
            clean_loom(loom)
        ] = index

    rows = []

    for row in range(
        2,
        worksheet.max_row + 1
    ):

        loom = clean_loom(
            worksheet.cell(
                row=row,
                column=loom_col
            ).value
        )

        week = parse_week(
            worksheet.cell(
                row=row,
                column=week_col
            ).value
        )

        production_date = parse_excel_date(
            worksheet.cell(
                row=row,
                column=date_col
            ).value
        )

        sequence = row_sequence.get(
            row,
            999999999
        )

        if production_date is None:

            date_sort = date.max

        else:

            date_sort = production_date

        starting_week = loom_settings.get(
            loom,
            {}
        ).get(
            "starting_week",
            1
        )

        if week is None:

            week_rank = 999

        else:

            week_rank = (
                week
                - starting_week
            ) % 52

        rows.append({

            "row":
                row,

            "loom_rank":
                loom_order.get(
                    loom,
                    999
                ),

            "week_rank":
                week_rank,

            "date":
                date_sort,

            "sequence":
                sequence
        })

    # --------------------------------------------------------
    # Sort.
    # --------------------------------------------------------

    rows.sort(
        key=lambda x: (
            x["loom_rank"],
            x["week_rank"],
            x["date"],
            x["sequence"],
            x["row"]
        )
    )

    # --------------------------------------------------------
    # Store complete row information.
    # --------------------------------------------------------

    max_col = worksheet.max_column

    row_data = {}

    for item in rows:

        original_row = item["row"]

        row_data[
            original_row
        ] = []

        for col in range(
            1,
            max_col + 1
        ):

            cell = worksheet.cell(
                row=original_row,
                column=col
            )

            row_data[
                original_row
            ].append({

                "value":
                    cell.value,

                "style":
                    copy(cell._style),

                "number_format":
                    cell.number_format,

                "font":
                    copy(cell.font),

                "fill":
                    copy(cell.fill),

                "border":
                    copy(cell.border),

                "alignment":
                    copy(cell.alignment),

                "protection":
                    copy(cell.protection),

                "hyperlink":
                    cell.hyperlink,

                "comment":
                    cell.comment
            })

    # --------------------------------------------------------
    # Rewrite.
    # --------------------------------------------------------

    for new_row, item in enumerate(
        rows,
        start=2
    ):

        original_row = item["row"]

        for col in range(
            1,
            max_col + 1
        ):

            target = worksheet.cell(
                row=new_row,
                column=col
            )

            if isinstance(
                target,
                MergedCell
            ):

                continue

            source = row_data[
                original_row
            ][col - 1]

            target.value = (
                source["value"]
            )

            target._style = copy(
                source["style"]
            )

            target.number_format = (
                source["number_format"]
            )

            target.font = copy(
                source["font"]
            )

            target.fill = copy(
                source["fill"]
            )

            target.border = copy(
                source["border"]
            )

            target.alignment = copy(
                source["alignment"]
            )

            target.protection = copy(
                source["protection"]
            )

            target.hyperlink = (
                source["hyperlink"]
            )

            target.comment = (
                source["comment"]
            )

    # --------------------------------------------------------
    # Restore merges.
    # --------------------------------------------------------

    for rng in merged_ranges:

        try:

            worksheet.merge_cells(
                rng
            )

        except Exception:

            pass


# ============================================================
# UI
# ============================================================

st.title(
    "🏭 Loom Production Planner"
)

st.write(
    "Plan production using weekly loom capacity, "
    "External Document sequence and production dates."
)


# ============================================================
# UPLOAD
# ============================================================

st.subheader(
    "1️⃣ Upload Production Excel"
)

uploaded_file = st.file_uploader(
    "Select Excel file",
    type=["xlsx"]
)


# ============================================================
# LOOM SETTINGS
# ============================================================

st.subheader(
    "2️⃣ Loom Settings"
)

st.caption(
    "Enable/disable looms and edit their weekly capacity "
    "and starting production week."
)

edited_settings = st.data_editor(

    st.session_state.loom_settings,

    use_container_width=True,

    hide_index=True,

    column_config={

        "Use":
            st.column_config.CheckboxColumn(
                "Use"
            ),

        "Loom":
            st.column_config.TextColumn(
                "Loom",
                disabled=True
            ),

        "Weekly Capacity":
            st.column_config.NumberColumn(
                "Weekly Capacity",
                min_value=1,
                step=50
            ),

        "Starting Week":
            st.column_config.NumberColumn(
                "Starting Week",
                min_value=1,
                max_value=52,
                step=1
            )
    },

    key="loom_settings_editor"
)


# ============================================================
# RESET
# ============================================================

if st.button(
    "🔄 Reset Loom Settings"
):

    st.session_state.loom_settings = (
        DEFAULT_SETTINGS.copy()
    )

    st.rerun()


# ============================================================
# VALIDATE BUTTON
# ============================================================

st.divider()

validate_button = st.button(
    "🔍 Check Existing Allocations",
    use_container_width=True
)


# ============================================================
# VALIDATION
# ============================================================

if validate_button:

    if uploaded_file is None:

        st.error(
            "Please upload the Excel file first."
        )

        st.stop()

    try:

        validation_workbook = load_workbook(
            filename=BytesIO(
                uploaded_file.getvalue()
            )
        )

        validation_sheet = (
            validation_workbook[
                validation_workbook.sheetnames[0]
            ]
        )

    except Exception as e:

        st.error(
            f"Unable to read Excel file: {e}"
        )

        st.stop()


    validation_settings = {}

    for _, row in edited_settings.iterrows():

        loom = clean_loom(
            row["Loom"]
        )

        validation_settings[
            loom
        ] = {

            "use":
                bool(row["Use"]),

            "capacity":
                float(row["Weekly Capacity"]),

            "starting_week":
                int(row["Starting Week"])
        }


    errors, warnings = (
        validate_input_allocations(

            validation_sheet,

            validation_settings,

            2,  # External Document No.

            6,  # Quantity

            7,  # Loom

            8,  # Production Week

            9   # Production Date
        )
    )


    if errors:

        st.error(
            f"❌ {len(errors)} error(s) found."
        )

        st.write(
            "Correct these issues before running "
            "production planning:"
        )

        for error in errors:

            st.write(
                f"🔴 {error}"
            )

    else:

        st.success(
            "✅ Week and date validation passed."
        )


    if warnings:

        st.warning(
            f"⚠️ {len(warnings)} warning(s) found."
        )

        for warning in warnings:

            st.write(
                f"🟡 {warning}"
            )

    elif not errors:

        st.success(
            "🎉 No warnings found."
        )


# ============================================================
# RUN PLANNING
# ============================================================

run_button = st.button(
    "🚀 Run Production Planning",
    type="primary",
    use_container_width=True
)


# ============================================================
# MAIN PROCESS
# ============================================================

if run_button:

    if uploaded_file is None:

        st.error(
            "Please upload the Excel file first."
        )

        st.stop()


    # ========================================================
    # LOOM SETTINGS
    # ========================================================

    loom_settings = {}

    for _, row in edited_settings.iterrows():

        loom = clean_loom(
            row["Loom"]
        )

        capacity = float(
            row["Weekly Capacity"]
        )

        starting_week = int(
            row["Starting Week"]
        )

        if capacity <= 0:

            st.error(
                f"{loom}: weekly capacity must be "
                "greater than 0."
            )

            st.stop()

        if not (
            1 <= starting_week <= 52
        ):

            st.error(
                f"{loom}: starting week must "
                "be between 1 and 52."
            )

            st.stop()

        loom_settings[
            loom
        ] = {

            "use":
                bool(row["Use"]),

            "capacity":
                capacity,

            "starting_week":
                starting_week
        }


    # ========================================================
    # LOAD WORKBOOK
    # ========================================================

    try:

        workbook = load_workbook(
            filename=BytesIO(
                uploaded_file.getvalue()
            )
        )

    except Exception as e:

        st.error(
            f"Unable to open Excel: {e}"
        )

        st.stop()


    worksheet = workbook[
        workbook.sheetnames[0]
    ]


    # ========================================================
    # FIXED INPUT COLUMNS
    # ========================================================
    #
    # A = No.
    # B = External Document No.
    # C = Description
    # D = Roll Width
    # E = Roll Length
    # F = Quantity
    # G = Loom
    #
    # H = Production Week
    # I = Production Date
    # J = Day
    #
    # ========================================================

    DOCUMENT_COL = 2
    QUANTITY_COL = 6
    LOOM_COL = 7

    WEEK_COL = 8
    DATE_COL = 9
    DAY_COL = 10


    # ========================================================
    # CHECK EXISTING WEEK / DATE DATA BEFORE PLANNING
    # ========================================================

    validation_errors, validation_warnings = (
        validate_input_allocations(

            worksheet,

            loom_settings,

            DOCUMENT_COL,

            QUANTITY_COL,

            LOOM_COL,

            WEEK_COL,

            DATE_COL
        )
    )


    if validation_errors:

        st.error(
            "❌ Planning stopped."
        )

        st.write(
            "The Excel file contains allocation errors. "
            "Fix them and upload the file again."
        )

        for error in validation_errors:

            st.write(
                f"🔴 {error}"
            )

        st.stop()


    # ========================================================
    # SHOW WARNINGS
    # ========================================================

    if validation_warnings:

        st.warning(
            "⚠️ Warnings found. Planning will continue."
        )

        for warning in validation_warnings:

            st.write(
                f"🟡 {warning}"
            )


    # ========================================================
    # CREATE OUTPUT COLUMNS
    # ========================================================

    # --------------------------------------------------------
    # H - Production Week
    # --------------------------------------------------------

    h_header = worksheet.cell(
        row=1,
        column=WEEK_COL
    ).value

    if (

        h_header is not None
        and
        str(h_header).strip().lower()
        != "production week"

    ):

        h_has_data = any(

            worksheet.cell(
                row=r,
                column=WEEK_COL
            ).value is not None

            for r in range(
                1,
                worksheet.max_row + 1
            )
        )

        if h_has_data:

            worksheet.insert_cols(
                WEEK_COL
            )


    worksheet.cell(
        row=1,
        column=WEEK_COL
    ).value = "Production Week"


    # --------------------------------------------------------
    # I - Production Date
    # --------------------------------------------------------

    i_header = worksheet.cell(
        row=1,
        column=DATE_COL
    ).value

    if (

        i_header is not None
        and
        str(i_header).strip().lower()
        != "production date"

    ):

        i_has_data = any(

            worksheet.cell(
                row=r,
                column=DATE_COL
            ).value is not None

            for r in range(
                1,
                worksheet.max_row + 1
            )
        )

        if i_has_data:

            worksheet.insert_cols(
                DATE_COL
            )


    worksheet.cell(
        row=1,
        column=DATE_COL
    ).value = "Production Date"


    # --------------------------------------------------------
    # J - Day
    # --------------------------------------------------------

    j_header = worksheet.cell(
        row=1,
        column=DAY_COL
    ).value

    if (

        j_header is not None
        and
        str(j_header).strip().lower()
        != "day"

    ):

        j_has_data = any(

            worksheet.cell(
                row=r,
                column=DAY_COL
            ).value is not None

            for r in range(
                1,
                worksheet.max_row + 1
            )
        )

        if j_has_data:

            worksheet.insert_cols(
                DAY_COL
            )


    worksheet.cell(
        row=1,
        column=DAY_COL
    ).value = "Day"


    # ========================================================
    # CLEAR OLD OUTPUT
    # ========================================================

    for row in range(
        2,
        worksheet.max_row + 1
    ):

        worksheet.cell(
            row=row,
            column=WEEK_COL
        ).value = None

        worksheet.cell(
            row=row,
            column=DATE_COL
        ).value = None

        worksheet.cell(
            row=row,
            column=DAY_COL
        ).value = None


    # ========================================================
    # GROUP DOCUMENTS BY LOOM
    # ========================================================

    loom_documents = {}

    for loom, settings in loom_settings.items():

        if not settings["use"]:
            continue

        loom_documents[
            loom
        ] = {}

        for row in range(
            2,
            worksheet.max_row + 1
        ):

            excel_loom = clean_loom(
                worksheet.cell(
                    row=row,
                    column=LOOM_COL
                ).value
            )

            if excel_loom != loom:
                continue

            document = clean_document(
                worksheet.cell(
                    row=row,
                    column=DOCUMENT_COL
                ).value
            )

            if document == "":

                document = (
                    f"ROW-{row}"
                )

            if document not in (
                loom_documents[loom]
            ):

                loom_documents[
                    loom
                ][document] = []

            loom_documents[
                loom
            ][document].append(
                row
            )


    # ========================================================
    # FIRST PASS:
    # ASSIGN PRODUCTION WEEKS
    # ========================================================

    row_week = {}

    row_sequence = {}

    loom_summary = {}


    for loom, settings in loom_settings.items():

        if not settings["use"]:
            continue


        weekly_capacity = (
            settings["capacity"]
        )

        current_week = (
            settings["starting_week"]
        )

        current_load = 0.0

        sequence = 0

        rows_processed = 0

        weeks_used = []

        documents = loom_documents.get(
            loom,
            {}
        )


        # ----------------------------------------------------
        # Process documents in original order.
        # ----------------------------------------------------

        for document, document_rows in (
            documents.items()
        ):

            for row in document_rows:

                quantity = parse_quantity(
                    worksheet.cell(
                        row=row,
                        column=QUANTITY_COL
                    ).value
                )

                if quantity is None:
                    continue


                # ------------------------------------------------
                # Weekly capacity.
                #
                # Individual product rows are NOT split.
                # ------------------------------------------------

                if (

                    current_load > 0

                    and

                    current_load + quantity
                    > weekly_capacity

                ):

                    current_week = (
                        next_week(
                            current_week
                        )
                    )

                    current_load = 0.0


                # ------------------------------------------------
                # Warn if one product is larger than capacity.
                # ------------------------------------------------

                if quantity > weekly_capacity:

                    st.warning(

                        f"{loom} / Document {document}: "
                        f"quantity {quantity:g} is greater "
                        f"than weekly capacity "
                        f"{weekly_capacity:g}. "
                        "The product will remain unsplit."
                    )


                # ------------------------------------------------
                # Assign week.
                # ------------------------------------------------

                row_week[
                    row
                ] = current_week


                # ------------------------------------------------
                # Preserve sequence.
                # ------------------------------------------------

                row_sequence[
                    row
                ] = sequence

                sequence += 1


                # ------------------------------------------------
                # Update load.
                # ------------------------------------------------

                current_load += quantity

                rows_processed += 1


                if current_week not in weeks_used:

                    weeks_used.append(
                        current_week
                    )


        # --------------------------------------------------------
        # Save summary.
        # --------------------------------------------------------

        loom_summary[
            loom
        ] = {

            "capacity":
                weekly_capacity,

            "starting_week":
                settings["starting_week"],

            "final_week":
                current_week,

            "rows_processed":
                rows_processed,

            "weeks_used":
                weeks_used
        }


    # ========================================================
    # SECOND PASS:
    # ASSIGN DATES
    # ========================================================

    row_date = {}


    for loom, settings in loom_settings.items():

        if not settings["use"]:
            continue


        starting_week = (
            settings["starting_week"]
        )


        # ----------------------------------------------------
        # Group rows by production week.
        # ----------------------------------------------------

        week_rows = {}


        for row, week in row_week.items():

            excel_loom = clean_loom(
                worksheet.cell(
                    row=row,
                    column=LOOM_COL
                ).value
            )

            if excel_loom != loom:
                continue


            quantity = parse_quantity(
                worksheet.cell(
                    row=row,
                    column=QUANTITY_COL
                ).value
            )

            if quantity is None:
                continue


            if week not in week_rows:

                week_rows[
                    week
                ] = []


            week_rows[
                week
            ].append(

                (
                    row,
                    quantity,
                    row_sequence.get(
                        row,
                        999999
                    )
                )
            )


        # ----------------------------------------------------
        # Process every production week.
        # ----------------------------------------------------

        for production_week, rows in (
            week_rows.items()
        ):

            rows.sort(
                key=lambda x: x[2]
            )


            # ------------------------------------------------
            # Calculate week starting date.
            # ------------------------------------------------

            year = 2026

            week_start = (
                get_week_start_date(
                    production_week,
                    year
                )
            )


            # ------------------------------------------------
            # Balance products across the week.
            # ------------------------------------------------

            assignments = (
                assign_balanced_dates(
                    rows,
                    week_start
                )
            )


            for row, assigned_date in (
                assignments.items()
            ):

                row_date[
                    row
                ] = assigned_date


    # ========================================================
    # WRITE RESULTS
    # ========================================================

    for row, week in row_week.items():

        if row not in row_date:
            continue


        production_date = (
            row_date[row]
        )


        # ----------------------------------------------------
        # Production Week
        # ----------------------------------------------------

        worksheet.cell(
            row=row,
            column=WEEK_COL
        ).value = (
            f"WK-{week}"
        )


        # ----------------------------------------------------
        # Production Date
        # ----------------------------------------------------

        worksheet.cell(
            row=row,
            column=DATE_COL
        ).value = (
            production_date
        )

        worksheet.cell(
            row=row,
            column=DATE_COL
        ).number_format = (
            "DD-MM-YYYY"
        )


        # ----------------------------------------------------
        # Day
        # ----------------------------------------------------

        worksheet.cell(
            row=row,
            column=DAY_COL
        ).value = (
            production_date.strftime(
                "%A"
            )
        )


    # ========================================================
    # SORT
    # ========================================================

    try:

        sort_planned_rows(

            worksheet,

            loom_settings,

            LOOM_COL,

            WEEK_COL,

            DATE_COL,

            DOCUMENT_COL,

            row_sequence
        )

    except Exception as e:

        st.warning(
            "Planning completed, but final sorting "
            f"could not be completed: {e}"
        )


    # ========================================================
    # SAVE
    # ========================================================

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)


    # ========================================================
    # SUCCESS
    # ========================================================

    st.success(
        "✅ Production planning completed successfully!"
    )


    # ========================================================
    # SUMMARY
    # ========================================================

    st.subheader(
        "📊 Planning Summary"
    )


    summary_rows = []


    for loom, info in (
        loom_summary.items()
    ):

        if info[
            "rows_processed"
        ] <= 0:

            continue


        summary_rows.append({

            "Loom":
                loom,

            "Weekly Capacity":
                info["capacity"],

            "Starting Week":
                f"WK-{info['starting_week']}",

            "Final Week":
                f"WK-{info['final_week']}",

            "Weeks Used":
                len(
                    info["weeks_used"]
                ),

            "Rows Processed":
                info["rows_processed"]
        })


    if summary_rows:

        st.dataframe(

            pd.DataFrame(
                summary_rows
            ),

            use_container_width=True,

            hide_index=True
        )

    else:

        st.warning(
            "No production rows were found for "
            "the selected/enabled looms."
        )


    # ========================================================
    # DOWNLOAD
    # ========================================================

    st.subheader(
        "4️⃣ Download Result"
    )


    st.download_button(

        label=(
            "📥 Download Planned Excel"
        ),

        data=output,

        file_name=(
            "Loom_wise_Production_Planning_Automated.xlsx"
        ),

        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),

        use_container_width=True
    )
